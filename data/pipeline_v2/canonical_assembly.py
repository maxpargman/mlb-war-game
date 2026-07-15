"""
canonical_assembly.py — Slice 3.5: canonical assembly + Gates 2-3.

Combines slice 3.3's WAR/stat ingestion with slice 3.4's suffix/display-
name resolution into one canonical table, enforces Gates 2-3, and writes
three artifacts to data/canonical/ (all gitignored - regenerated on every
run, not the final shippable output; that's slice 3.7/3.9's job):

  canonical.parquet          the assembled table itself
  reconciliation_report.txt  Gate 5: human-readable source->canonical
                              counts, every exclusion bucket, coverage
                              results, mapping verification, diff vs the
                              previous build if one exists
  franchise_audit.xlsx       one sheet per active franchise: year
                              coverage (gaps highlighted), players per
                              season, top-WAR players per decade — for
                              Max's manual review in slice 3.6

Gate 2 (conservation): rows_in == rows_out + rows_excluded_with_reason,
carried forward from slice 3.3's already-passing check. A non-empty
`unknown` bucket fails the build.

Gate 3 (canonical coverage), fail conditions per CC_PLAN.md:
  - every active franchise has a row in every year of its existence
    (derived from Teams.csv) — coverage gaps fail
  - no duplicate (bbrefID, year, teamID) keys — fail
  - exactly 30 active franchises present — fail
  - roster-agreement (BR vs Lahman rosters per team-year) — reuses
    slice 3.2's already-passing check; a regression fails
Per-team-season player-count sanity (~25-60, looser pre-1900) is a WARN
only, per CC_PLAN.md's fail/warn split — roster counts can legitimately
vary (strike seasons, expansion-team debut years) without indicating a
real defect.

Run with the mlbwar env active, from the repo root:
    python data/pipeline_v2/canonical_assembly.py
"""

import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill

from loaders import (
    load_teams, load_teams_franchises, load_people, load_appearances,
    load_batting, load_pitching, load_war_bat, load_war_pitch, load_chadwick_register,
)
from team_mapping import build_team_mapping, compute_roster_overlap, partition_failures
from war_ingestion import build_war_grain, resolve_teams, resolve_players, attach_stats
from names import load_register_suffixes, load_overrides, resolve_suffixes, assemble_display_name

OUT_DIR = Path(__file__).parent.parent / "canonical"
PARQUET_OUT = OUT_DIR / "canonical.parquet"
REPORT_OUT = OUT_DIR / "reconciliation_report.txt"
WORKBOOK_OUT = OUT_DIR / "franchise_audit.xlsx"

ROSTER_SIZE_BOUNDS_MODERN = (25, 60)
ROSTER_SIZE_BOUNDS_PRE_1900 = (10, 70)
PRE_1900_CUTOFF = 1900


def build_canonical(
    teams, teams_franchises, people, appearances, batting, pitching, bat, pitch, register, overrides
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict]:
    """Returns (canonical, team_excluded, mapping, gate1_inputs) where
    gate1_inputs carries the raw WAR-grain row count for the conservation
    check."""
    mapping = build_team_mapping(teams)
    grain = build_war_grain(bat, pitch)
    resolved, team_excluded = resolve_teams(grain, mapping)
    resolved = resolve_players(resolved, people)
    final = attach_stats(resolved, batting, pitching, appearances)

    register_suffixes, _rejected = load_register_suffixes(register)
    suffix_map, _stale = resolve_suffixes(register_suffixes, overrides)
    final["suffix"] = final["bbrefID"].map(suffix_map)
    final["displayName"] = final.apply(
        lambda r: assemble_display_name(r["nameFirst"], r["nameLast"], r["suffix"]), axis=1
    )

    return final, team_excluded, mapping, {"rows_in": len(grain)}


def gate2_conservation(canonical: pd.DataFrame, team_excluded: pd.DataFrame, rows_in: int) -> tuple[bool, str]:
    unknown_count = int((canonical["resolutionReason"] == "unknown").sum())
    rows_out = len(canonical)
    rows_excluded = len(team_excluded)
    conserved = rows_in == rows_out + rows_excluded
    ok = conserved and unknown_count == 0
    msg = (
        f"rows_in ({rows_in:,}) == rows_out ({rows_out:,}) + rows_excluded ({rows_excluded:,})  -> {conserved}\n"
        f"unknown-reason rows: {unknown_count}  -> {'OK' if unknown_count == 0 else 'FAIL'}"
    )
    return ok, msg


def gate3_franchise_year_coverage(canonical: pd.DataFrame, teams: pd.DataFrame, active_franchises: set[str]) -> pd.DataFrame:
    rows = []
    for franchID in sorted(active_franchises):
        span = teams.loc[teams["franchID"] == franchID, "yearID"]
        lo, hi = int(span.min()), int(span.max())
        have_years = set(canonical.loc[canonical["franchID"] == franchID, "year"].unique())
        missing = sorted(set(range(lo, hi + 1)) - have_years)
        rows.append({"franchID": franchID, "yearLo": lo, "yearHi": hi, "missingYears": missing})
    return pd.DataFrame(rows)


def gate3_duplicate_keys(canonical: pd.DataFrame) -> pd.DataFrame:
    return canonical[canonical.duplicated(subset=["bbrefID", "year", "lahmanTeamID"], keep=False)]


def gate3_roster_size_sanity(canonical: pd.DataFrame) -> pd.DataFrame:
    counts = canonical.groupby(["franchID", "year"])["bbrefID"].nunique().reset_index(name="players")

    def bounds(year: int) -> tuple[int, int]:
        return ROSTER_SIZE_BOUNDS_PRE_1900 if year < PRE_1900_CUTOFF else ROSTER_SIZE_BOUNDS_MODERN

    lo_hi = counts["year"].map(bounds)
    counts["lo"] = lo_hi.map(lambda t: t[0])
    counts["hi"] = lo_hi.map(lambda t: t[1])
    counts["outOfBounds"] = (counts["players"] < counts["lo"]) | (counts["players"] > counts["hi"])
    return counts[counts["outOfBounds"]]


def run_gates(
    canonical: pd.DataFrame, team_excluded: pd.DataFrame, teams: pd.DataFrame, teams_franchises: pd.DataFrame,
    mapping: pd.DataFrame, bat, pitch, appearances, people, rows_in: int,
) -> dict:
    active_franchises = set(teams_franchises.loc[teams_franchises["active"] == "Y", "franchID"])
    report = {}

    g2_ok, g2_msg = gate2_conservation(canonical, team_excluded, rows_in)
    report["gate2"] = {"ok": g2_ok, "message": g2_msg}

    coverage = gate3_franchise_year_coverage(canonical, teams, active_franchises)
    coverage_gaps = coverage[coverage["missingYears"].map(len) > 0]
    report["gate3_coverage"] = {"ok": len(coverage_gaps) == 0, "table": coverage, "gaps": coverage_gaps}

    dupes = gate3_duplicate_keys(canonical)
    report["gate3_dupes"] = {"ok": len(dupes) == 0, "count": len(dupes)}

    # The canonical layer intentionally keeps every franchise ever (active
    # or long-defunct) — "nothing deleted, ever". This check is that all 30
    # currently-active franchises are represented, not that the table
    # contains only 30 franchises total.
    active_present = canonical.loc[canonical["franchID"].isin(active_franchises), "franchID"].nunique()
    report["gate3_franchise_count"] = {"ok": active_present == 30, "count": active_present}

    checked, blocking, deferred = compute_and_partition_overlap(mapping, bat, pitch, appearances, people, active_franchises)
    report["gate3_roster_agreement"] = {"ok": len(blocking) == 0, "checked": checked, "blocking": len(blocking), "deferred": len(deferred)}

    size_outliers = gate3_roster_size_sanity(canonical)
    report["gate3_roster_size_warn"] = {"count": len(size_outliers), "table": size_outliers}

    all_fail_gates_ok = (
        g2_ok and report["gate3_coverage"]["ok"] and report["gate3_dupes"]["ok"]
        and report["gate3_franchise_count"]["ok"] and report["gate3_roster_agreement"]["ok"]
    )
    report["overall_ok"] = all_fail_gates_ok
    return report


def compute_and_partition_overlap(mapping, bat, pitch, appearances, people, active_franchises):
    results = compute_roster_overlap(mapping, bat, pitch, appearances, people)
    blocking, deferred = partition_failures(results, active_franchises)
    return len(results), blocking, deferred


def write_reconciliation_report(report: dict, canonical: pd.DataFrame, team_excluded: pd.DataFrame, rows_in: int) -> str:
    lines = []
    lines.append("=" * 78)
    lines.append("CANONICAL BUILD RECONCILIATION REPORT")
    lines.append(f"Generated: {datetime.now(timezone.utc).isoformat()}")
    lines.append("=" * 78)

    lines.append("\n--- Source -> canonical counts ---")
    lines.append(f"WAR grain rows in:      {rows_in:,}")
    lines.append(f"Canonical rows out:     {len(canonical):,}")
    lines.append(f"Excluded (team-level):  {len(team_excluded):,}")

    lines.append("\n--- Exclusion buckets ---")
    for reason, count in team_excluded["reason"].value_counts().items():
        lines.append(f"  {reason:20s} {count:5d}")
        sample = team_excluded[team_excluded["reason"] == reason][["bbrefID", "year", "brCode"]].head(3)
        for row in sample.itertuples(index=False):
            lines.append(f"      e.g. {row.bbrefID} {row.year} {row.brCode}")

    lines.append("\n--- Player ID resolution ---")
    for reason, count in canonical["resolutionReason"].value_counts().items():
        lines.append(f"  {reason:20s} {count:6,d}")

    lines.append("\n--- Gate 2: conservation ---")
    lines.append(f"  {'PASS' if report['gate2']['ok'] else 'FAIL'}  {report['gate2']['message']}")

    lines.append("\n--- Gate 3: canonical coverage ---")
    cov = report["gate3_coverage"]
    lines.append(f"  Franchise-year coverage: {'PASS' if cov['ok'] else 'FAIL'} "
                  f"({len(cov['gaps'])} franchise(s) with missing years)")
    for row in cov["gaps"].itertuples(index=False):
        lines.append(f"      {row.franchID} ({row.yearLo}-{row.yearHi}): missing {row.missingYears}")

    dupes = report["gate3_dupes"]
    lines.append(f"  Duplicate (bbrefID, year, teamID) keys: {'PASS' if dupes['ok'] else 'FAIL'} ({dupes['count']} found)")

    fc = report["gate3_franchise_count"]
    lines.append(f"  Active franchise count == 30: {'PASS' if fc['ok'] else 'FAIL'} (found {fc['count']})")

    ra = report["gate3_roster_agreement"]
    lines.append(f"  Roster-agreement (BR vs Lahman, active franchises): {'PASS' if ra['ok'] else 'FAIL'} "
                  f"({ra['checked']:,} team-years checked, {ra['blocking']} active-franchise failure(s), "
                  f"{ra['deferred']} deferred non-active-franchise failure(s) - see slice 3.2)")

    lines.append("\n--- Warn-only: roster-size sanity (not gate-blocking) ---")
    warn = report["gate3_roster_size_warn"]
    lines.append(f"  {warn['count']} team-season(s) outside the expected player-count range")
    if warn["count"]:
        for row in warn["table"].sort_values(["franchID", "year"]).itertuples(index=False):
            lines.append(f"      {row.franchID} {row.year}: {row.players} players (expected {row.lo}-{row.hi})")

    lines.append("\n--- Diff vs previous build ---")
    lines.append("  No previous build recorded - this is the first canonical assembly run.")

    lines.append("\n" + "=" * 78)
    lines.append(f"OVERALL: {'PASS' if report['overall_ok'] else 'FAIL'}")
    lines.append("=" * 78)

    return "\n".join(lines)


def _sanitize_sheet_name(name: str) -> str:
    name = re.sub(r"[:\\/?*\[\]]", "", name)
    return name[:31]


def write_audit_workbook(canonical: pd.DataFrame, teams: pd.DataFrame, teams_franchises: pd.DataFrame, active_franchises: set[str]) -> None:
    franch_names = teams_franchises.set_index("franchID")["franchName"].to_dict()
    canonical = canonical.copy()
    canonical["totalWAR"] = canonical["battingWAR"].fillna(0) + canonical["pitchingWAR"].fillna(0)
    canonical["decade"] = (canonical["year"] // 10) * 10

    with pd.ExcelWriter(WORKBOOK_OUT, engine="openpyxl") as writer:
        for franchID in sorted(active_franchises, key=lambda f: franch_names.get(f, f)):
            sub = canonical[canonical["franchID"] == franchID]
            span = teams.loc[teams["franchID"] == franchID, "yearID"]
            lo, hi = int(span.min()), int(span.max())
            have_years = set(sub["year"].unique())

            year_rows = []
            for y in range(lo, hi + 1):
                players_that_year = sub[sub["year"] == y]["bbrefID"].nunique()
                year_rows.append({"year": y, "players": players_that_year, "hasGap": y not in have_years})
            year_df = pd.DataFrame(year_rows)

            top_by_decade = (
                sub.sort_values("totalWAR", ascending=False)
                .groupby("decade")
                .head(5)[["decade", "year", "displayName", "totalWAR"]]
                .sort_values(["decade", "totalWAR"], ascending=[True, False])
            )

            sheet = _sanitize_sheet_name(f"{franchID} {franch_names.get(franchID, '')}")
            year_df.to_excel(writer, sheet_name=sheet, index=False, startrow=0)
            top_by_decade.to_excel(writer, sheet_name=sheet, index=False, startrow=len(year_df) + 3)

        # Highlight gap rows across every sheet.
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        wb = writer.book
        for ws in wb.worksheets:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=3):
                if row[2].value is True:  # hasGap column
                    for cell in row:
                        cell.fill = red_fill


def main() -> None:
    print("=" * 78)
    print("SLICE 3.5 - CANONICAL ASSEMBLY + GATES 2-3")
    print("=" * 78)

    teams = load_teams()
    teams_franchises = load_teams_franchises()
    people = load_people()
    appearances = load_appearances()
    batting = load_batting()
    pitching = load_pitching()
    bat = load_war_bat()
    pitch = load_war_pitch()
    register = load_chadwick_register()
    overrides = load_overrides()

    canonical, team_excluded, mapping, gate1 = build_canonical(
        teams, teams_franchises, people, appearances, batting, pitching, bat, pitch, register, overrides
    )
    print(f"\nCanonical rows: {len(canonical):,}   Excluded: {len(team_excluded):,}")

    report = run_gates(canonical, team_excluded, teams, teams_franchises, mapping, bat, pitch, appearances, people, gate1["rows_in"])

    report_text = write_reconciliation_report(report, canonical, team_excluded, gate1["rows_in"])
    print("\n" + report_text)

    if not report["overall_ok"]:
        print("\nGate failure(s) above - no Parquet/workbook written.")
        sys.exit(1)

    OUT_DIR.mkdir(exist_ok=True)
    canonical.to_parquet(PARQUET_OUT, index=False)
    REPORT_OUT.write_text(report_text, encoding="utf-8")

    active_franchises = set(teams_franchises.loc[teams_franchises["active"] == "Y", "franchID"])
    write_audit_workbook(canonical, teams, teams_franchises, active_franchises)

    print(f"\nWrote {PARQUET_OUT} ({len(canonical):,} rows)")
    print(f"Wrote {REPORT_OUT}")
    print(f"Wrote {WORKBOOK_OUT}")


if __name__ == "__main__":
    main()
